import sys
import os
import pandas as pd
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(SCRIPT_DIR)))
sys.path.append(REPO_ROOT)
from Step_1_Process_Macro_Flows_and_Balance_Demand import scenario_names

from lcoe.liquid_fuels.ethanol_upgrade_lcoe.a_json_to_csv import ASSET_FILES

TEMPLATE_XLSX_PATH = os.path.join(SCRIPT_DIR, "LCOE_ETHANOL_UPGRADE_TEMPLATE.xlsx")

CSV_TO_XLSX_MAP = {
    "id":                  "id",
    "commodity":           "commodity",
    "elec_consumption":    "elec_consumption (MWh/MWh-ethanol)",
    "elec_production":     "elec_production (MWh/MWh-ethanol)",
    "h2_consumption":    "h2_consumption (MWh/MWh-ethanol)",
    "gasoline_production": "gasoline_production (MWh-gasoline/MWh-ethanol)",
    "diesel_production": "diesel_production (MWh-diesel/MWh-ethanol)",
    "jetfuel_production": "jetfuel_production (MWh-jetfuel/MWh-ethanol)",
    "capture_rate":        "process_capture_rate (t-CO2/MWh-ethanol)",
    "emission_rate":       "process_emission_rate (t-CO2/MWh-ethanol)",
    "investment_cost":     "investment_cost ($/yr per MW ethanol)",
    "fixed_om_cost":       "fixed_om_cost ($/yr per MW ethanol)",
    "variable_om_cost":    "variable_om_cost ($/MWh-ethanol)",
}

HEADER_ROW = 3
DATA_START_ROW = 4

for label in scenario_names:
    out_dir = os.path.join(SCRIPT_DIR, label)

    all_rows = []
    for _, csv_name, _ in ASSET_FILES:
        csv_path = os.path.join(out_dir, csv_name)
        if not os.path.exists(csv_path):
            print(f"Warning: CSV not found for scenario {label}: {csv_path}")
            continue
        all_rows.append(pd.read_csv(csv_path))

    if not all_rows:
        print(f"Skipping scenario {label}: no CSVs found (run a_json_to_csv.py first)")
        continue

    combined_df = pd.concat(all_rows, ignore_index=True)

    wb = load_workbook(TEMPLATE_XLSX_PATH)
    ws = wb["lc_detailed"]

    xlsx_header_to_idx = {}
    for cell in ws[HEADER_ROW]:
        if cell.value is not None:
            xlsx_header_to_idx[cell.value] = cell.column

    csv_field_to_col_idx = {}
    for csv_field, xlsx_header in CSV_TO_XLSX_MAP.items():
        if xlsx_header in xlsx_header_to_idx:
            csv_field_to_col_idx[csv_field] = xlsx_header_to_idx[xlsx_header]

    # Clear any stale template values in the columns we're about to repopulate,
    # so leftover sample rows beyond the new data don't carry over.
    for row in range(DATA_START_ROW, ws.max_row + 1):
        for col_idx in csv_field_to_col_idx.values():
            ws.cell(row=row, column=col_idx).value = None

    for r_offset, row_data in enumerate(combined_df.itertuples(index=False)):
        for csv_field, col_idx in csv_field_to_col_idx.items():
            value = getattr(row_data, csv_field, None)
            if pd.isna(value):
                value = None
            ws.cell(row=DATA_START_ROW + r_offset, column=col_idx, value=value)

    out_xlsx_path = os.path.join(SCRIPT_DIR, f"LCOE_ETHANOL_UPGRADE_{label}.xlsx")
    wb.save(out_xlsx_path)
    print(f"Scenario {label}: {len(combined_df)} rows written to {out_xlsx_path}")

print("DONE CSV POPULATED INTO XLSX FOR ALL SCENARIOS")