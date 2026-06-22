import pandas as pd
from openpyxl import load_workbook

# change this manual file path!
manual_file_path = "/Users/abbie/MacroEnergy-Abbie.jl/MacroEnergyExamples/9Zone_US/allethylene_168hr/results_006"

# ── Paths ──────────────────────────────────────────────────────────────────────
DUALS_CSV     = f'{manual_file_path}/results/balance_duals.csv'
CO2_DUALS_CSV = f'{manual_file_path}/results/co2_cap_duals.csv'
XLSX_PATH     = "/Users/abbie/MacroEnergy-Abbie.jl/MacroEnergyExamples/lcoe_plots/bio_ethanol_lcoe/LCOE_BIOETHANOL.xlsx"

# ── Mapping: xlsx dual column header → CSV column prefix ──────────────────────
DUAL_COLUMN_MAP = {
    "elec_demand ($/MWh)":     "elec",
    "h2_balance ($/MWh)":      "h2",
    "h2_demand ($/MWh)":       "h2_demand",
    "ng_balance ($/MWh)":      "natgas",
    "ng_demand ($/MWh)":       "natgas_demand",
    "ethanol_balance ($/MWh)": "ethanol",
    "ethanol_demand ($/MWh)":  "ethanol_demand",
    "ethylene_balance ($/t)": "ethylene",
    "ethylene_demand ($/t)":  "ethylene_demand",
    "gasoline_balance ($/MWh)":"gasoline",
    "gasoline_demand ($/MWh)": "gasoline_demand",
    "diesel_balance ($/MWh)":  "diesel",
    "diesel_demand ($/MWh)":   "diesel_demand",
    "jetfuel_balance ($/MWh)": "jetfuel",
    "jetfuel_demand ($/MWh)":  "jetfuel_demand",
    "bioherb_supply ($/t)":  "bioherb",
    "biowood_supply ($/t)":  "biowood",
    "bioagri_supply ($/t)":  "bioagri",
    "biocorn_supply ($/t)":  "corn",
    "co2_captured ($/t)":  "co2_captured",
}

# ── co2_sink comes from a separate CSV ────────────────────────────────────────
co2_df = pd.read_csv(CO2_DUALS_CSV)
co2_sink_value = co2_df.loc[co2_df["Node"] == "co2_sink", "CO2_Shadow_Price"].values
CO2_SINK_VALUE = round(float(co2_sink_value[0]), 6) if len(co2_sink_value) > 0 else None

HEADER_ROW     = 3
DATA_START_ROW = 4

# ── Load data ──────────────────────────────────────────────────────────────────
duals_df = pd.read_csv(DUALS_CSV)
wb = load_workbook(XLSX_PATH)
ws = wb.active

# ── Build column index map from xlsx header row ────────────────────────────────
col_name_to_idx = {}
for cell in ws[HEADER_ROW]:
    if cell.value:
        col_name_to_idx[cell.value] = cell.column

# ── Helper: extract zone from asset id ────────────────────────────────────────
def extract_zone(asset_id: str) -> str:
    return asset_id.split("_")[0]

# ── Clear any existing dual values before writing ─────────────────────────────
all_dual_cols = set(DUAL_COLUMN_MAP.keys()) | {"co2_sink ($/t)"}
dual_col_indices = {
    col_name_to_idx[col]
    for col in all_dual_cols
    if col in col_name_to_idx
}

row = DATA_START_ROW
while True:
    asset_id = ws.cell(row=row, column=col_name_to_idx.get("id", 2)).value
    if asset_id is None:
        break
    for col_idx in dual_col_indices:
        ws.cell(row=row, column=col_idx).value = None
    row += 1

# ── Iterate over data rows ─────────────────────────────────────────────────────
row = DATA_START_ROW
while True:
    asset_id = ws.cell(row=row, column=col_name_to_idx.get("id", 2)).value
    if asset_id is None:
        break

    zone = extract_zone(str(asset_id))

    for xlsx_col, csv_prefix in DUAL_COLUMN_MAP.items():
        if xlsx_col not in col_name_to_idx:
            continue
        csv_col = f"{csv_prefix}_{zone}"
        if csv_col not in duals_df.columns:
            continue
        avg_val = duals_df[csv_col].mean()
        ws.cell(row=row, column=col_name_to_idx[xlsx_col], value=round(avg_val, 6))

    if "co2_sink ($/t)" in col_name_to_idx and CO2_SINK_VALUE is not None:
        ws.cell(row=row, column=col_name_to_idx["co2_sink ($/t)"], value=CO2_SINK_VALUE)

    row += 1

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save(XLSX_PATH)
print(f"Done! Populated duals for {row - DATA_START_ROW} rows.")
print(f"  co2_sink value used: {CO2_SINK_VALUE}")
print("DONE DUALS POPULATED INTO XLSX")