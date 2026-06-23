import sys
import os
import pandas as pd
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
sys.path.append(REPO_ROOT)
from Step_1_Process_Macro_Flows_and_Balance_Demand import scenario_names

from a_json_to_csv import ASSET_FILES

TEMPLATE_XLSX_PATH = os.path.join(SCRIPT_DIR, "LCOE_BIOETHANOL_TEMPLATE.xlsx")

CSV_TO_XLSX_MAP = {
    "id":                  "id",
    "commodity":           "commodity",
    "elec_consumption":    "elec_consumption (MWh/t-bio)",
    "elec_production":     "elec_production (MWh/t-bio)",
    "ethanol_production": "ethanol_production (MWh-ethanol/t-bio)",
    "natgas_consumption":  "natgas_consumption (MWh/t-bio)",
    "co2_biomass_content":  "co2_biomass_content (t-CO2/t-bio)",
    "process_capture_rate":        "process_capture_rate (t-CO2/t-bio)",
    "process_emission_rate":       "process_emission_rate (t-CO2/t-bio)",
    "fuel_capture_rate":        "fuel_capture_rate (t-CO2/t-bio)",
    "fuel_emission_rate":       "fuel_emission_rate (t-CO2/t-bio)",
    "investment_cost":     "investment_cost ($/yr per t-bio/hr)",
    "fixed_om_cost":       "fixed_om_cost ($/yr per t-bio/hr)",
    "variable_om_cost":    "variable_om_cost ($/t-bio)",
}

HEADER_ROW = 3
DATA_START_ROW = 4

for label in scenario_names:
    out_dir = os.path.join(SCRIPT_DIR, label)

    all_rows = []
    for _, csv_name, _ in ASSET_FILES:
        csv_path = os.path.join(out_dir, csv_name)
        if not os.path.exists(csv_path):
            print(f"Warning: CSV not found for scenario {label}: {csv_path}")
            continue
        all_rows.append(pd.read_csv(csv_path))

    if not all_rows:
        print(f"Skipping scenario {label}: no CSVs found (run a_json_to_csv.py first)")
        continue

    combined_df = pd.concat(all_rows, ignore_index=True)

    wb = load_workbook(TEMPLATE_XLSX_PATH)
    ws = wb["lc_detailed"]

    xlsx_header_to_idx = {}
    for cell in ws[HEADER_ROW]:
        if cell.value is not None:
            xlsx_header_to_idx[cell.value] = cell.column

    csv_field_to_col_idx = {}
    for csv_field, xlsx_header in CSV_TO_XLSX_MAP.items():
        if xlsx_header in xlsx_header_to_idx:
            csv_field_to_col_idx[csv_field] = xlsx_header_to_idx[xlsx_header]

    for r_offset, row_data in enumerate(combined_df.itertuples(index=False)):
        for csv_field, col_idx in csv_field_to_col_idx.items():
            value = getattr(row_data, csv_field, None)
            if pd.isna(value):
                value = None
            ws.cell(row=DATA_START_ROW + r_offset, column=col_idx, value=value)

    out_xlsx_path = os.path.join(SCRIPT_DIR, f"LCOE_BIOETHANOL_{label}.xlsx")
    wb.save(out_xlsx_path)
    print(f"Scenario {label}: {len(combined_df)} rows written to {out_xlsx_path}")

print("DONE CSV POPULATED INTO XLSX FOR ALL SCENARIOS")
