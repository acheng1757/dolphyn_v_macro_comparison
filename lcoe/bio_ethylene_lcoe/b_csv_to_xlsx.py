import sys
import json
import pandas as pd
from openpyxl import load_workbook

sys.path.append("/Users/abbie/MacroEnergy-Abbie.jl")
from MacroEnergyExamples.lcoe_plots.bio_ethylene_lcoe.a_json_to_csv import json_to_csv_transforms

# change this manual file path!
manual_file_path = "/Users/abbie/MacroEnergy-Abbie.jl/MacroEnergyExamples/9Zone_US/allethylene_168hr"

ASSETS_PATH = f'{manual_file_path}/assets/'
XLSX_PATH = "/Users/abbie/MacroEnergy-Abbie.jl/MacroEnergyExamples/lcoe_plots/bio_ethylene_lcoe/LCOE_BIO_Ethylene.xlsx" # this will stay the same

json_files = [
    "ethanol_dehydration.json"
]

CSV_TO_XLSX_MAP = {
    "id":                  "id",
    "h2_consumption":      "h2_consumption (MWh/MWh-ethanol)",
    "elec_consumption":    "elec_consumption (MWh/MWh-ethanol)",
    "ethylene_production": "ethylene_production (t-ethylene/MWh-ethanol)",
    "natgas_consumption":  "natgas_consumption (MWh/MWh-ethanol)",
    "capture_rate":        "capture_rate (t-CO2/MWh-ethanol)",
    "emission_rate":       "emission_rate (t-CO2/MWh-ethanol)",
    "investment_cost":     "investment_cost ($/yr per MW ethanol)",
    "fixed_om_cost":       "fixed_om_cost ($/yr per MW ethanol)",
    "variable_om_cost":    "variable_om_cost ($/MWh-ethanol)",
}

FIELDS = list(CSV_TO_XLSX_MAP.keys())

all_rows = []
for json_file in json_files:
    with open(ASSETS_PATH + json_file) as f:
        data = json.load(f)
    rows = json_to_csv_transforms(data)
    all_rows.extend(rows)

combined_df = pd.DataFrame(all_rows)

# Duplicate all rows, appending a suffix to the id of each duplicate
originals = combined_df.to_dict(orient="records")
duplicates = []
for row in originals:
    duplicate = row.copy()
    duplicate["id"] = f"{row['id']}_ethanol_biochemical"
    duplicates.append(duplicate)

combined_df = pd.DataFrame(originals + duplicates)

wb = load_workbook(XLSX_PATH)
ws = wb.active

HEADER_ROW = 2
DATA_START_ROW = 3

xlsx_header_to_idx = {}
for cell in ws[HEADER_ROW]:
    if cell.value is not None:
        xlsx_header_to_idx[cell.value] = cell.column

csv_field_to_col_idx = {}
for csv_field, xlsx_header in CSV_TO_XLSX_MAP.items():
    if xlsx_header in xlsx_header_to_idx:
        csv_field_to_col_idx[csv_field] = xlsx_header_to_idx[xlsx_header]

for r_offset, row_data in enumerate(combined_df.itertuples(index=False)):
    for csv_field, col_idx in csv_field_to_col_idx.items():
        value = getattr(row_data, csv_field, None)
        ws.cell(row=DATA_START_ROW + r_offset, column=col_idx, value=value)

wb.save(XLSX_PATH)
print(f"Done! {len(combined_df)} rows written.")
print("DONE CSV POPULATED INTO XLSX")