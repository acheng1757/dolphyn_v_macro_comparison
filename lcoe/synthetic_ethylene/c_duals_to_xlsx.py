import os
import sys
import pandas as pd
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
sys.path.append(REPO_ROOT)
from Step_1_Process_Macro_Flows_and_Balance_Demand import macro_base_dir, scenario_names, macro_scenario_paths

# ── Mapping: xlsx dual column header → CSV column prefix ──────────────────────
DUAL_COLUMN_MAP = {
    "elec_demand ($/MWh)":     "elec",
    "h2_balance ($/MWh)":      "h2_",
    "ng_balance ($/MWh)":      "natgas",
    "gasoline_balance ($/MWh)":"gasoline",
    "co2_captured ($/t)":  "co2_captured",
}

HEADER_ROW     = 3
DATA_START_ROW = 4


def extract_zone(asset_id: str) -> str:
    return asset_id.split("_")[0]


def load_scenario_duals(label):
    """
    Returns (duals_df, co2_sink) for a scenario, or (None, None) if
    balance_duals.csv isn't found. duals_df is the raw balance_duals.csv
    frame (one column per zone/global balance); use get_dual_value() to
    look up a specific xlsx dual column for a given zone. If
    co2_cap_duals.csv is missing (e.g. the scenario has no CO2 cap),
    co2_sink defaults to 0 instead of bailing out.
    """
    results_dir = os.path.join(macro_base_dir, macro_scenario_paths[label])
    duals_csv_path = os.path.join(results_dir, "balance_duals.csv")
    co2_duals_csv_path = os.path.join(results_dir, "co2_cap_duals.csv")

    if not os.path.exists(duals_csv_path):
        return None, None

    if os.path.exists(co2_duals_csv_path):
        co2_df = pd.read_csv(co2_duals_csv_path)
        co2_sink_value = co2_df.loc[co2_df["Node"] == "co2_sink", "CO2_Shadow_Price"].values
        co2_sink = round(float(co2_sink_value[0]), 6) if len(co2_sink_value) > 0 else None
    else:
        co2_sink = 0.0

    duals_df = pd.read_csv(duals_csv_path)
    return duals_df, co2_sink


def load_scenario_capacity(label):
    """
    Returns the capacity.csv DataFrame for a scenario (one row per
    resource_id with its net built `capacity`), or None if not found.
    """
    results_dir = os.path.join(macro_base_dir, macro_scenario_paths[label])
    capacity_csv_path = os.path.join(results_dir, "capacity.csv")
    if not os.path.exists(capacity_csv_path):
        return None
    return pd.read_csv(capacity_csv_path)


def get_dual_value(duals_df, xlsx_col, zone):
    """
    Look up the dual for one xlsx column ("elec_demand ($/MWh)", etc.) in
    one zone. Most balances are zoned (e.g. "elec_CA"); some have no zone
    breakdown and instead post a single national dual (e.g.
    "gasoline_global", "ethylene_demand_global") that applies regardless
    of zone. Returns None if neither variant exists in duals_df.
    """
    csv_prefix = DUAL_COLUMN_MAP.get(xlsx_col)
    if csv_prefix is None:
        return None
    zone_col = f"{csv_prefix}_{zone}"
    global_col = f"{csv_prefix}_global"
    if zone_col in duals_df.columns:
        return round(duals_df[zone_col].mean(), 6)
    elif global_col in duals_df.columns:
        return round(duals_df[global_col].mean(), 6)
    return None


if __name__ == "__main__":
    for label in scenario_names:
        xlsx_path = os.path.join(SCRIPT_DIR, f"LCOE_SYNTHETIC_{label}.xlsx")

        duals_df, co2_sink = load_scenario_duals(label)
        if duals_df is None:
            print(f"Warning: balance_duals.csv/co2_cap_duals.csv not found for scenario {label}")
            continue
        if not os.path.exists(xlsx_path):
            print(f"Warning: {xlsx_path} not found for scenario {label} (run b_csv_to_xlsx.py first)")
            continue

        wb = load_workbook(xlsx_path)
        ws = wb["lc_detailed"]

        # ── Build column index map from xlsx header row ───────────────────────
        col_name_to_idx = {}
        for cell in ws[HEADER_ROW]:
            if cell.value:
                col_name_to_idx[cell.value] = cell.column

        # ── Clear any existing dual values before writing ─────────────────────
        all_dual_cols = set(DUAL_COLUMN_MAP.keys()) | {"co2_sink ($/t)"}
        dual_col_indices = {
            col_name_to_idx[col]
            for col in all_dual_cols
            if col in col_name_to_idx
        }

        for row in range(DATA_START_ROW, ws.max_row + 1):
            for col_idx in dual_col_indices:
                ws.cell(row=row, column=col_idx).value = None

        # ── Iterate over data rows ──────────────────────────────────────────────
        row = DATA_START_ROW
        while True:
            asset_id = ws.cell(row=row, column=col_name_to_idx.get("id", 2)).value
            if asset_id is None:
                break

            zone = extract_zone(str(asset_id))

            for xlsx_col in DUAL_COLUMN_MAP:
                if xlsx_col not in col_name_to_idx:
                    continue
                value = get_dual_value(duals_df, xlsx_col, zone)
                if value is None:
                    continue
                ws.cell(row=row, column=col_name_to_idx[xlsx_col], value=value)

            if "co2_sink ($/t)" in col_name_to_idx and co2_sink is not None:
                ws.cell(row=row, column=col_name_to_idx["co2_sink ($/t)"], value=co2_sink)

            row += 1

        wb.save(xlsx_path)
        print(f"Scenario {label}: populated duals for {row - DATA_START_ROW} rows in {xlsx_path} (co2_sink={co2_sink})")

    print("DONE DUALS POPULATED INTO XLSX FOR ALL SCENARIOS")
