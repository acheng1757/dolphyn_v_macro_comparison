import sys
import json
import pandas as pd
from openpyxl import load_workbook

# change this manual file path!
manual_file_path = "/Users/abbie/MacroEnergy-Abbie.jl/MacroEnergyExamples/9Zone_US/allethylene_168hr"

sys.path.append("/Users/abbie/MacroEnergy-Abbie.jl")
from MacroEnergyExamples.lcoe_plots.sc_esc_ethylene_lcoe.a_json_to_csv_ETHYLENE import json_to_csv_transforms

ASSETS_PATH = f'{manual_file_path}/assets/'
XLSX_PATH = "/Users/abbie/MacroEnergy-Abbie.jl/MacroEnergyExamples/lcoe_plots/sc_esc_ethylene_lcoe/LCOE_SC_ESC_Ethylene.xlsx"

json_files = [
    "thermalsteamcracker.json",
    "thermalsteamcracker_retrofit_option.json",
    "electrifiedsteamcracker_retrofit_option.json",
]

CSV_TO_XLSX_MAP = {
    "id":                  "id",
    "commodity":           "commodity",
    "h2_consumption":      "h2_consumption (MWh/MWh-ethane)",
    "h2_production":       "h2_production (MWh/MWh-ethane)",
    "elec_consumption":    "elec_consumption (MWh/MWh-ethane)",
    "ethylene_production": "ethylene_production (t-ethylene/MWh-ethane)",
    "natgas_consumption":  "natgas_consumption (MWh/MWh-ethane)",
    "natgas_production":   "natgas_production (MWh/MWh-ethane)",
    "capture_rate":        "ASSET_capture_rate (t-CO2/MWh-ethane)",
    "emission_rate":       "ASSET_emission_rate (t-CO2/MWh-ethane)",
    "investment_cost":     "investment_cost ($/yr per MW ethane)",
    "fixed_om_cost":       "fixed_om_cost ($/yr per MW ethane)",
    "variable_om_cost":    "variable_om_cost ($/MWh-ethane)",
}

FIELDS = list(CSV_TO_XLSX_MAP.keys())

all_rows = []
for json_file in json_files:
    with open(ASSETS_PATH + json_file) as f:
        data = json.load(f)
    rows = json_to_csv_transforms(data)
    all_rows.extend(rows)

combined_df = pd.DataFrame(all_rows)

wb = load_workbook(XLSX_PATH)
ws = wb.active

HEADER_ROW = 2
DATA_START_ROW = 3

xlsx_header_to_idx = {}
for cell in ws[HEADER_ROW]:
    if cell.value is not None:
        xlsx_header_to_idx[cell.value] = cell.column

csv_field_to_col_idx = {}
for csv_field, xlsx_header in CSV_TO_XLSX_MAP.items():
    if xlsx_header in xlsx_header_to_idx:
        csv_field_to_col_idx[csv_field] = xlsx_header_to_idx[xlsx_header]

ccs_col_idx = xlsx_header_to_idx.get("CCS")

# ---- ADDED: clear all relevant columns from DATA_START_ROW to last used row ----
cols_to_clear = set(csv_field_to_col_idx.values())
if ccs_col_idx is not None:
    cols_to_clear.add(ccs_col_idx)

for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
    for cell in row:
        if cell.column in cols_to_clear:
            cell.value = None
# ---------------------------------------------------------------------------------

for r_offset, row_data in enumerate(combined_df.itertuples(index=False)):
    for csv_field, col_idx in csv_field_to_col_idx.items():
        value = getattr(row_data, csv_field, None)
        ws.cell(row=DATA_START_ROW + r_offset, column=col_idx, value=value)

    if ccs_col_idx is not None:
        row_id = getattr(row_data, "id", "")
        ccs_value = 0.9 if "CC90" in str(row_id) else 0
        ws.cell(row=DATA_START_ROW + r_offset, column=ccs_col_idx, value=ccs_value)

wb.save(XLSX_PATH)
print(f"Done! {len(combined_df)} rows written.")
print("DONE CSV POPULATED INTO XLSX")